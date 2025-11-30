import os
import re
from typing import List, Any, Dict
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn
import openpyxl
import xlrd

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Configurações ---
UPLOAD_DIR = "arquivos_recebidos"
OUTPUT_DIR = "banco_de_dados_clientes"
TEMPLATE_XLSX = "modelo.xlsx"
TEMPLATE_XLS = "modelo.xls"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

class DadosPlanilha(BaseModel):
    dados: List[List[Any]]

# --- Índices baseados no seu Frontend (index.html) ---
IDX_NOME = 1          # Nome Unidade
IDX_CNPJ = 38         # CNPJ Unidade
IDX_CPF = 109         # CPF Unidade
IDX_CAEPF = 110       # CAEPF Unidade
IDX_CNO = 113         # CNO Unidade (Ajuste se necessário conforme sua lista exata)

def sanitizar_valor(valor) -> str:
    """Remove pontuação e espaços extras de documentos e nomes."""
    if not valor:
        return ""
    # Converte para string e remove caracteres que não sejam letras ou números
    s = str(valor).strip()
    return re.sub(r'[^\w\s]', '', s) # Mantém letras, números e espaços

def gerar_id_unico(linha: List[Any]) -> str:
    """
    Cria um ID único para o arquivo baseado na hierarquia:
    CNPJ > CNO > CAEPF > CPF > Nome
    """
    # Extrai valores das colunas (com segurança caso a linha seja curta)
    nome = sanitizar_valor(linha[IDX_NOME]) if len(linha) > IDX_NOME else "SemNome"
    cnpj = sanitizar_valor(linha[IDX_CNPJ]) if len(linha) > IDX_CNPJ else ""
    cno = sanitizar_valor(linha[IDX_CNO]) if len(linha) > IDX_CNO else ""
    caepf = sanitizar_valor(linha[IDX_CAEPF]) if len(linha) > IDX_CAEPF else ""
    cpf = sanitizar_valor(linha[IDX_CPF]) if len(linha) > IDX_CPF else ""

    # Lógica de Prioridade para o Prefixo do Arquivo
    prefixo = ""
    if cnpj:
        prefixo = cnpj # CNPJ completo diferencia Filial (0002) de Matriz (0001)
    elif cno:
        prefixo = "CNO_" + cno
    elif caepf:
        prefixo = "CAEPF_" + caepf
    elif cpf:
        prefixo = "CPF_" + cpf
    else:
        # Se não tiver documento nenhum, usa um hash simples ou apenas o nome
        prefixo = "SEM_DOC"

    # Monta o nome final: DOC_NOME.xlsx (Substitui espaços do nome por _)
    nome_arquivo = f"{prefixo}_{nome.replace(' ', '_')}"
    
    # Limita tamanho para não dar erro no Windows (max 255 chars) e garante segurança
    return nome_arquivo[:100]

def carregar_template():
    if os.path.exists(TEMPLATE_XLSX):
        return openpyxl.load_workbook(TEMPLATE_XLSX)
    elif os.path.exists(TEMPLATE_XLS):
        # Conversão rápida de .xls antigo
        rb = xlrd.open_workbook(TEMPLATE_XLS)
        sheet_old = rb.sheet_by_index(0)
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "Modelo 1"
        for r in range(sheet_old.nrows):
            ws_new.append(sheet_old.row_values(r))
        return wb_new
    else:
        raise FileNotFoundError("Modelo não encontrado.")

@app.post("/gerar-excel/")
def processar_dados_clientes(payload: DadosPlanilha):
    try:
        if not payload.dados:
            return {"status": "erro", "mensagem": "Vazio"}

        # 1. Agrupar linhas pelo ID ÚNICO (Documento + Nome)
        grupos: Dict[str, List[List[Any]]] = {}

        for linha in payload.dados:
            id_unico = gerar_id_unico(linha)
            
            if id_unico not in grupos:
                grupos[id_unico] = []
            grupos[id_unico].append(linha)

        arquivos_gerados = []

        # 2. Salvar cada grupo em seu arquivo
        for id_arquivo, linhas in grupos.items():
            nome_arquivo = f"{id_arquivo}.xlsx"
            caminho = os.path.join(OUTPUT_DIR, nome_arquivo)

            # Abre existente ou cria novo
            if os.path.exists(caminho):
                wb = openpyxl.load_workbook(caminho)
                ws = wb.active
            else:
                wb = carregar_template()
                ws = wb.active

            # Adiciona dados
            for row in linhas:
                ws.append(row)

            wb.save(caminho)
            arquivos_gerados.append(nome_arquivo)

        return {
            "status": "ok",
            "mensagem": f"Sucesso! {len(arquivos_gerados)} arquivos processados.",
            "detalhes": arquivos_gerados
        }

    except Exception as e:
        print(f"Erro: {e}")
        return {"status": "erro", "mensagem": str(e)}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)